"""Loader for the historical ``SellUp Stock Data.xlsx`` mapping sheet.

This is the record of links confirmed by hand before the tool existed. It is a
flat three-column sheet:

    POS Stock Type ID | SellUp Variation ID | SellUp Variation Name

with two sentinels in the first column:

* ``-`` / blank -- placeholder row, no link recorded
* ``not in pos`` -- the SellUp listing has been reviewed and has no POS source

One SellUp SKU is frequently fed by several POS IDs. In the real data
``SKU-000074155`` is linked to three rows at once::

    31628  Used  17 PRO MAX 256GB     SILVER  -> Excellent      (col K)
    31242  New   17 PRO MAX 256GB NA  SILVER  -> Not Activated  (col G)
    31243  New   17 PRO MAX 256GB A   SILVER  -> Activated      (col I)

so links are grouped by **condition slot** and summed within a slot. They are
never summed across slots -- that would report one physical handset as
available in three different conditions at once.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import IO

import openpyxl

from .normalize import clean

SENTINEL_NOT_IN_POS = "not in pos"
SENTINEL_BLANK = {"", "-", "none", "n/a", "na"}

EXPECTED_HEADERS = ("POS Stock Type ID", "SellUp Variation ID", "SellUp Variation Name")


class SeedParseError(Exception):
    """Raised when the seed mapping sheet has an unexpected layout."""


@dataclass
class SeedMapping:
    """Confirmed POS-to-SellUp links carried over from previous runs."""

    links: dict[str, list[str]] = field(default_factory=dict)
    not_in_pos: set[str] = field(default_factory=set)
    names: dict[str, str] = field(default_factory=dict)

    rows_read: int = 0
    placeholder_rows: int = 0

    def pos_ids(self) -> set[str]:
        return {pid for ids in self.links.values() for pid in ids}

    def sku_for_pos_id(self) -> dict[str, list[str]]:
        """Reverse index: POS ID -> the SellUp SKUs it feeds."""
        out: dict[str, list[str]] = {}
        for sku, ids in self.links.items():
            for pid in ids:
                out.setdefault(pid, []).append(sku)
        return out

    def shared_pos_ids(self) -> dict[str, list[str]]:
        """POS IDs feeding more than one SellUp SKU -- a double-count risk."""
        return {k: v for k, v in self.sku_for_pos_id().items() if len(set(v)) > 1}


def _normalise_pos_id(value: object) -> str | None:
    """Return a clean POS ID, or ``None`` for a blank/placeholder cell.

    Excel stores these IDs as floats, so ``31628.0`` has to become ``"31628"``.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return str(int(value))
    text = clean(value)
    if text.lower() in SENTINEL_BLANK:
        return None
    return text


def _is_not_in_pos(value: object) -> bool:
    return isinstance(value, str) and clean(value).lower() == SENTINEL_NOT_IN_POS


def load_seed_mapping(source: str | IO[bytes]) -> SeedMapping:
    """Read ``SellUp Stock Data.xlsx`` into a :class:`SeedMapping`."""
    workbook = openpyxl.load_workbook(source, data_only=True, read_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        mapping = SeedMapping()

        rows = worksheet.iter_rows(min_row=1, max_col=3, values_only=True)
        try:
            header = next(rows)
        except StopIteration as exc:
            raise SeedParseError("Seed mapping sheet is empty.") from exc

        actual = [clean(h).upper() for h in header[:3]]
        expected = [h.upper() for h in EXPECTED_HEADERS]
        if actual != expected:
            raise SeedParseError(
                "Seed mapping sheet header mismatch.\n"
                f"Expected: {list(EXPECTED_HEADERS)}\nFound:    {actual}"
            )

        seen: set[tuple[str, str]] = set()

        for pos_raw, sku_raw, name_raw in rows:
            sku = clean(sku_raw)
            if not sku or sku.lower() in SENTINEL_BLANK:
                mapping.placeholder_rows += 1
                continue

            mapping.rows_read += 1
            name = clean(name_raw)
            if name and name.lower() not in SENTINEL_BLANK:
                mapping.names.setdefault(sku, name)

            if _is_not_in_pos(pos_raw):
                mapping.not_in_pos.add(sku)
                continue

            pos_id = _normalise_pos_id(pos_raw)
            if pos_id is None:
                mapping.placeholder_rows += 1
                continue

            if (sku, pos_id) in seen:
                continue
            seen.add((sku, pos_id))
            mapping.links.setdefault(sku, []).append(pos_id)

        # A SKU that ended up with a real link is no longer "not in pos".
        mapping.not_in_pos -= set(mapping.links)

        return mapping
    finally:
        workbook.close()


def summarise_seed(mapping: SeedMapping) -> dict[str, int]:
    """Counts for the dashboard."""
    return {
        "linked_skus": len(mapping.links),
        "linked_pairs": sum(len(v) for v in mapping.links.values()),
        "distinct_pos_ids": len(mapping.pos_ids()),
        "multi_source_skus": sum(1 for v in mapping.links.values() if len(v) > 1),
        "shared_pos_ids": len(mapping.shared_pos_ids()),
        "not_in_pos": len(mapping.not_in_pos),
        "placeholder_rows": mapping.placeholder_rows,
    }
