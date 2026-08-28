"""Reader for the Mister Mobile POS masterlist export (``stock_report*.xlsx``).

Hard rules enforced here:

* **Column F only.** ``Available Quantity`` already nets off reserved and
  transit stock. Column G is never read, and per-branch columns are never summed.
* **TELCO rows are dropped.** SellUp has no telco listings and the PRIMARY /
  TELCO pools are never combined.
* **Export sets and freebies are dropped.**
* **Apple and Used rows are kept.** SellUp trades in both.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from typing import IO

import openpyxl

from . import config
from .normalize import DeviceSpec, clean, colour_key, maker, parse_pos_model, upper


class PosParseError(Exception):
    """Raised when the POS export does not have the expected shape."""


@dataclass
class PosRow:
    """A single sellable POS masterlist row."""

    stock_type_id: str
    category: str            # 'New' | 'Used'
    brand: str
    model: str
    colour: str
    available_qty: int       # Column F, verbatim
    spec: DeviceSpec
    excel_row: int

    @property
    def maker(self) -> str:
        return maker(self.brand)

    @property
    def is_used(self) -> bool:
        return self.category.upper() == "USED"

    @property
    def slot(self) -> str:
        """Which SellUp quantity column this row feeds."""
        if self.is_used:
            return config.SLOT_USED_EXCELLENT
        if self.spec.activation == config.ACTIVATION_ACTIVATED:
            return config.SLOT_NEW_A
        return config.SLOT_NEW_NA

    @property
    def label(self) -> str:
        return f"{self.stock_type_id}:{self.model}|{self.colour}"

    def match_key(self) -> tuple:
        return (self.maker, *self.spec.identity(), colour_key(self.colour), self.slot)

    def loose_key(self) -> tuple:
        return (self.maker, *self.spec.loose_identity(), colour_key(self.colour), self.slot)


@dataclass
class PosExclusion:
    """A POS row that was deliberately not considered for SellUp."""

    stock_type_id: str
    brand: str
    model: str
    colour: str
    available_qty: int
    reason: str


@dataclass
class PosMasterlist:
    """Parsed POS export: the sellable rows plus an audit trail of exclusions."""

    rows: list[PosRow]
    exclusions: list[PosExclusion]
    total_rows_read: int

    def by_id(self) -> dict[str, PosRow]:
        return {r.stock_type_id: r for r in self.rows}

    def with_stock(self) -> list[PosRow]:
        return [r for r in self.rows if r.available_qty > 0]


def _is_freebie(model_upper: str) -> bool:
    return any(tok in model_upper for tok in config.FREEBIE_TOKENS)


def _is_export_set(model_upper: str) -> bool:
    """Detect a parallel-import region set."""
    work = model_upper
    for phrase in config.US_CHARGER_PHRASES:
        work = work.replace(phrase, " ")
    for token in config.EXPORT_TOKENS:
        if re.search(rf"\b{token}\b", work):
            return True
    return False


def _coerce_qty(value: object) -> int | None:
    """Column F to a non-negative int, or None when it is not a number."""
    if value is None or value == "":
        return None
    try:
        qty = int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None
    return max(qty, 0)


def validate_pos_workbook(worksheet) -> list[str]:
    """Return a list of human-readable problems with the POS sheet layout."""
    problems: list[str] = []

    headers = [
        clean(worksheet.cell(config.POS_HEADER_ROW, col).value)
        for col in range(1, len(config.POS_REQUIRED_HEADERS) + 1)
    ]
    for idx, expected in enumerate(config.POS_REQUIRED_HEADERS):
        actual = headers[idx] if idx < len(headers) else ""
        if actual.upper() != expected.upper():
            problems.append(
                f"POS row {config.POS_HEADER_ROW} column {idx + 1}: "
                f"expected header '{expected}', found '{actual or '(blank)'}'."
            )

    sub = clean(worksheet.cell(config.POS_SUBHEADER_ROW, config.POS_COL_AVAILABLE_QTY).value)
    if sub.upper() != "AVAILABLE QUANTITY":
        problems.append(
            f"POS cell F{config.POS_SUBHEADER_ROW} must read 'Available Quantity' "
            f"(found '{sub or '(blank)'}'). Column F is the only stock figure this "
            "tool will read, so the export layout has changed."
        )

    if worksheet.max_row < config.POS_FIRST_DATA_ROW:
        problems.append("POS export contains no data rows below the header.")

    return problems


def load_pos_masterlist(source: str | IO[bytes]) -> PosMasterlist:
    """Read a POS ``stock_report`` export into a :class:`PosMasterlist`."""
    workbook = openpyxl.load_workbook(source, data_only=True, read_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]

        problems = validate_pos_workbook(worksheet)
        if problems:
            raise PosParseError("\n".join(problems))

        rows: list[PosRow] = []
        exclusions: list[PosExclusion] = []
        total = 0

        for excel_row, cells in enumerate(
            worksheet.iter_rows(
                min_row=config.POS_FIRST_DATA_ROW,
                max_col=config.POS_COL_AVAILABLE_QTY,
                values_only=True,
            ),
            start=config.POS_FIRST_DATA_ROW,
        ):
            stock_type_id = clean(cells[config.POS_COL_STOCK_TYPE_ID - 1])
            category = clean(cells[config.POS_COL_CATEGORY - 1])
            brand = clean(cells[config.POS_COL_BRAND - 1])
            model = clean(cells[config.POS_COL_MODEL - 1])
            colour = clean(cells[config.POS_COL_COLOR - 1])
            raw_qty = cells[config.POS_COL_AVAILABLE_QTY - 1]

            if not stock_type_id and not model:
                continue

            total += 1
            model_upper = upper(model)

            qty = _coerce_qty(raw_qty)
            if qty is None:
                exclusions.append(
                    PosExclusion(stock_type_id, brand, model, colour, 0,
                                 "Column F (Available Quantity) is blank or non-numeric")
                )
                continue

            if _is_freebie(model_upper):
                exclusions.append(
                    PosExclusion(stock_type_id, brand, model, colour, qty,
                                 "Freebie / giveaway unit")
                )
                continue

            if _is_export_set(model_upper):
                exclusions.append(
                    PosExclusion(stock_type_id, brand, model, colour, qty,
                                 "Export / parallel-import set")
                )
                continue

            spec = parse_pos_model(model)

            if spec.channel in config.EXCLUDED_CHANNELS:
                exclusions.append(
                    PosExclusion(stock_type_id, brand, model, colour, qty,
                                 "TELCO channel stock (not sold on SellUp)")
                )
                continue

            rows.append(
                PosRow(
                    stock_type_id=stock_type_id,
                    category=category,
                    brand=brand,
                    model=model,
                    colour=colour,
                    available_qty=qty,
                    spec=spec,
                    excel_row=excel_row,
                )
            )

        return PosMasterlist(rows=rows, exclusions=exclusions, total_rows_read=total)
    finally:
        workbook.close()


def summarise_pos(masterlist: PosMasterlist) -> dict[str, int]:
    """Counts for the dashboard."""
    rows = masterlist.rows
    return {
        "rows_read": masterlist.total_rows_read,
        "sellable_rows": len(rows),
        "excluded_rows": len(masterlist.exclusions),
        "rows_with_stock": len(masterlist.with_stock()),
        "new_not_activated": sum(1 for r in rows if r.slot == config.SLOT_NEW_NA),
        "new_activated": sum(1 for r in rows if r.slot == config.SLOT_NEW_A),
        "used_excellent": sum(1 for r in rows if r.slot == config.SLOT_USED_EXCELLENT),
        "total_units": sum(r.available_qty for r in rows),
    }
