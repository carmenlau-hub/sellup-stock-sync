"""Read and write the SellUp SKU Registry workbook.

This is the file the reviewing actually happens in. The layout mirrors the
existing ``Shopee_Match_Review`` registry so both look and behave the same:

* navy ``1F3864`` headers with white bold text for index and decision columns
* orange ``F4B183`` for platform-side (SellUp) columns
* yellow ``FFD966`` for masterlist-side (POS) columns
* green ``C6E0B4`` for the cells the reviewer fills in
* 10pt body text, frozen header row, autofilter on the data tabs

The ``New Masterlist SKUs`` sheet carries the tool's ranked suggestions
alongside each unresolved row, so a decision can be made in Excel without
going back to the app.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from typing import IO, Iterable, Sequence

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from . import config
from .normalize import clean
from .pipeline import LockedMatch, NewMasterlistSku, PipelineResult


class RegistryParseError(Exception):
    """Raised when an uploaded registry is missing tabs or columns."""


# --------------------------------------------------------------------------
# Styling helpers
# --------------------------------------------------------------------------

_NAVY_FILL = PatternFill("solid", fgColor=config.COLOR_NAVY)
_ORANGE_FILL = PatternFill("solid", fgColor=config.COLOR_ORANGE)
_YELLOW_FILL = PatternFill("solid", fgColor=config.COLOR_YELLOW)
_GREEN_FILL = PatternFill("solid", fgColor=config.COLOR_GREEN)

_NAVY_FONT = Font(bold=True, size=config.BASE_FONT_SIZE, color=config.COLOR_WHITE)
_DARK_FONT = Font(bold=True, size=config.BASE_FONT_SIZE, color=config.COLOR_BLACK)
_BODY_FONT = Font(size=config.BASE_FONT_SIZE)

# Which fill each header gets. Anything unlisted falls back to navy.
_PLATFORM_HEADERS = {
    "SellUp Sheet",
    "SellUp SKU ID",
    "SellUp Model",
    "Storage",
    "Connectivity",
    "SellUp Colour",
    "Condition",
    "Current Seller Stock",
    "Suggested SellUp SKU ID",
    "Suggested SellUp Listing",
    "Confidence",
    "Score",
    "Why",
    "Alternative 2",
    "Alternative 3",
}
_MASTERLIST_HEADERS = {
    "LOCKED Masterlist ID(s)",
    "ML Category",
    "ML Model(s)|Color",
    "ML Available Qty",
    "Target Stock",
    "Masterlist Stock Type ID",
    "Category",
    "Brand",
    "Model",
    "Color",
    "Available Qty",
    "Corrected Masterlist ID",
}
_INPUT_HEADERS = set(config.NEW_SKUS_INPUT_COLUMNS)

_COLUMN_WIDTHS: dict[str, float] = {
    "#": 5,
    "SellUp Sheet": 14,
    "SellUp SKU ID": 17,
    "SellUp Model": 30,
    "Storage": 10,
    "Connectivity": 20,
    "SellUp Colour": 20,
    "Condition": 20,
    "LOCKED Masterlist ID(s)": 24,
    "ML Category": 11,
    "ML Model(s)|Color": 40,
    "ML Available Qty": 16,
    "Target Stock": 12,
    "# SKUs": 8,
    "How linked": 14,
    "Masterlist Stock Type ID": 22,
    "Category": 10,
    "Brand": 14,
    "Model": 34,
    "Color": 20,
    "Available Qty": 13,
    "Suggested SellUp SKU ID": 22,
    "Suggested SellUp Listing": 42,
    "Confidence": 12,
    "Score": 8,
    "Why": 40,
    "Alternative 2": 42,
    "Alternative 3": 42,
    "Link to SellUp SKU ID": 22,
    "Reviewer Decision": 22,
    "Notes": 30,
    "Current Seller Stock": 18,
    "Corrected Masterlist ID": 22,
}


def _fill_for(header: str) -> tuple[PatternFill, Font]:
    if header in _INPUT_HEADERS:
        return _GREEN_FILL, _DARK_FONT
    if header in _PLATFORM_HEADERS:
        return _ORANGE_FILL, _DARK_FONT
    if header in _MASTERLIST_HEADERS:
        return _YELLOW_FILL, _DARK_FONT
    return _NAVY_FILL, _NAVY_FONT


def _write_header(worksheet, headers: Sequence[str]) -> None:
    for idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(1, idx, header)
        fill, font = _fill_for(header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        worksheet.column_dimensions[get_column_letter(idx)].width = _COLUMN_WIDTHS.get(
            header, 16
        )
    worksheet.freeze_panes = "A2"


def _write_rows(worksheet, rows: Iterable[Sequence[object]], header_count: int) -> int:
    written = 0
    for offset, row in enumerate(rows, start=2):
        for idx, value in enumerate(row, start=1):
            cell = worksheet.cell(offset, idx, value)
            cell.font = _BODY_FONT
        written += 1
    if written:
        worksheet.auto_filter.ref = f"A1:{get_column_letter(header_count)}{written + 1}"
    return written


# --------------------------------------------------------------------------
# Reading an existing registry
# --------------------------------------------------------------------------

@dataclass
class Registry:
    """A parsed SKU registry: confirmed links plus previous classifications."""

    links: dict[str, list[str]] = field(default_factory=dict)
    not_selling: set[str] = field(default_factory=set)
    not_yet: set[str] = field(default_factory=set)
    no_pos_source: set[str] = field(default_factory=set)
    decisions: dict[str, dict] = field(default_factory=dict)

    def classified_pos_ids(self) -> set[str]:
        return self.not_selling | self.not_yet


def validate_registry_workbook(workbook) -> list[str]:
    """Check that every required tab is present and readable.

    Only the columns the reader actually depends on are enforced, so a
    registry that has been re-ordered or annotated in Excel still loads.
    """
    problems: list[str] = []

    required_columns = {
        config.SHEET_LOCKED: ("SellUp SKU ID", "LOCKED Masterlist ID(s)"),
        config.SHEET_NEW_SKUS: ("Masterlist Stock Type ID", "Reviewer Decision"),
        config.SHEET_MATCH_REVIEW: ("SellUp SKU ID",),
        config.SHEET_NOT_SELLING: ("Masterlist Stock Type ID",),
        config.SHEET_NOT_YET: ("Masterlist Stock Type ID",),
    }

    for sheet_name, columns in required_columns.items():
        if sheet_name not in workbook.sheetnames:
            problems.append(f"Registry is missing the '{sheet_name}' worksheet.")
            continue
        worksheet = workbook[sheet_name]
        present = {
            clean(worksheet.cell(1, i).value).upper()
            for i in range(1, worksheet.max_column + 1)
        }
        for column in columns:
            if column.upper() not in present:
                problems.append(
                    f"'{sheet_name}' is missing the '{column}' column. "
                    "Do not delete or rename the header row."
                )
    return problems


def _column_index(worksheet, header: str) -> int | None:
    for idx in range(1, worksheet.max_column + 1):
        if clean(worksheet.cell(1, idx).value).upper() == header.upper():
            return idx
    return None


def load_registry(source: str | IO[bytes]) -> Registry:
    """Read a previously exported registry back into memory."""
    workbook = openpyxl.load_workbook(source, data_only=True)
    try:
        problems = validate_registry_workbook(workbook)
        if problems:
            raise RegistryParseError("\n".join(problems))

        registry = Registry()

        # Link History -> the complete map, when the registry carries one.
        # Read first so it takes precedence over Locked Matches, which only
        # lists SKUs that had live POS stock on the day of export.
        if config.SHEET_LINK_HISTORY in workbook.sheetnames:
            worksheet = workbook[config.SHEET_LINK_HISTORY]
            sku_col = _column_index(worksheet, "SellUp SKU ID")
            ids_col = _column_index(worksheet, "LOCKED Masterlist ID(s)")
            status_col = _column_index(worksheet, "Status")
            if sku_col and ids_col:
                for row in range(2, worksheet.max_row + 1):
                    sku = clean(worksheet.cell(row, sku_col).value)
                    if not sku:
                        continue
                    status = (
                        clean(worksheet.cell(row, status_col).value).lower()
                        if status_col
                        else ""
                    )
                    if status == "no pos source":
                        registry.no_pos_source.add(sku)
                        continue
                    ids = clean(worksheet.cell(row, ids_col).value)
                    if not ids:
                        continue
                    bucket = registry.links.setdefault(sku, [])
                    for pid in (p.strip() for p in ids.split(",")):
                        if pid and pid not in bucket:
                            bucket.append(pid)

        # Locked Matches -> confirmed links.
        worksheet = workbook[config.SHEET_LOCKED]
        sku_col = _column_index(worksheet, "SellUp SKU ID")
        ids_col = _column_index(worksheet, "LOCKED Masterlist ID(s)")
        if sku_col and ids_col:
            for row in range(2, worksheet.max_row + 1):
                sku = clean(worksheet.cell(row, sku_col).value)
                ids = clean(worksheet.cell(row, ids_col).value)
                if not sku or not ids:
                    continue
                bucket = registry.links.setdefault(sku, [])
                for pid in (p.strip() for p in ids.split(",")):
                    if pid and pid not in bucket:
                        bucket.append(pid)

        # Classification tabs -> POS IDs already dealt with.
        for sheet_name, target in (
            (config.SHEET_NOT_SELLING, registry.not_selling),
            (config.SHEET_NOT_YET, registry.not_yet),
        ):
            worksheet = workbook[sheet_name]
            id_col = _column_index(worksheet, "Masterlist Stock Type ID")
            if not id_col:
                continue
            for row in range(2, worksheet.max_row + 1):
                pid = clean(worksheet.cell(row, id_col).value)
                if pid:
                    target.add(pid)
                    registry.decisions[pid] = {
                        "decision": (
                            config.DECISION_NOT_SELLING
                            if sheet_name == config.SHEET_NOT_SELLING
                            else config.DECISION_NOT_YET
                        ),
                        "linked_sku_id": "",
                        "notes": "carried over from registry",
                    }

        # New Masterlist SKUs -> the decisions made in Excel since last time.
        worksheet = workbook[config.SHEET_NEW_SKUS]
        id_col = _column_index(worksheet, "Masterlist Stock Type ID")
        link_col = _column_index(worksheet, "Link to SellUp SKU ID")
        dec_col = _column_index(worksheet, "Reviewer Decision")
        notes_col = _column_index(worksheet, "Notes")
        if id_col and dec_col:
            for row in range(2, worksheet.max_row + 1):
                pid = clean(worksheet.cell(row, id_col).value)
                if not pid:
                    continue
                decision = clean(worksheet.cell(row, dec_col).value)
                sku = clean(worksheet.cell(row, link_col).value) if link_col else ""
                notes = clean(worksheet.cell(row, notes_col).value) if notes_col else ""

                # Filling in a SKU is taken as a link even if the decision
                # column was left blank -- that is the common shorthand when
                # working quickly in Excel.
                if sku and decision not in config.TERMINAL_DECISIONS:
                    decision = config.DECISION_LINKED
                if decision not in config.TERMINAL_DECISIONS:
                    continue

                registry.decisions[pid] = {
                    "decision": decision,
                    "linked_sku_id": sku,
                    "notes": notes,
                }
                if decision == config.DECISION_LINKED and sku:
                    bucket = registry.links.setdefault(sku, [])
                    if pid not in bucket:
                        bucket.append(pid)

        # Match Review -> SellUp listings acknowledged as having no POS source.
        worksheet = workbook[config.SHEET_MATCH_REVIEW]
        sku_col = _column_index(worksheet, "SellUp SKU ID")
        if sku_col:
            for row in range(2, worksheet.max_row + 1):
                sku = clean(worksheet.cell(row, sku_col).value)
                if sku:
                    registry.no_pos_source.add(sku)

        return registry
    finally:
        workbook.close()


# --------------------------------------------------------------------------
# Writing a registry
# --------------------------------------------------------------------------

def _summary_sheet(workbook, result: PipelineResult, generated: str) -> None:
    worksheet = workbook.create_sheet(config.SHEET_SUMMARY, 0)
    worksheet.column_dimensions["A"].width = 2
    worksheet.column_dimensions["B"].width = 44
    worksheet.column_dimensions["C"].width = 12

    title = worksheet.cell(2, 2, "SellUp Stock Bulk Update — Match Registry")
    title.font = Font(bold=True, size=config.TITLE_FONT_SIZE, color=config.COLOR_NAVY)

    for idx, header in enumerate(("Category", "Count"), start=2):
        cell = worksheet.cell(4, idx, header)
        cell.fill = _NAVY_FILL
        cell.font = _NAVY_FONT

    metrics = result.metrics()
    rows = [
        ("Locked Matches (stock synced)", metrics["locked_updated"]),
        ("  of which linked automatically", metrics["auto_linked"]),
        ("Still to review in this file", metrics["requiring_review"]),
        ("Not Selling in SellUp", metrics["not_selling"]),
        ("Not on SellUp Yet", metrics["not_yet"]),
        ("Match Review (no POS source)", len(result.match_review)),
        ("Validation errors & warnings", metrics["validation_errors"]),
        ("Quantity cells written", metrics["cells_to_write"]),
        ("Listings delisted (set to 0)", metrics["delisted"]),
        ("Total units synced", metrics["units_synced"]),
    ]
    for offset, (label, value) in enumerate(rows, start=5):
        worksheet.cell(offset, 2, label).font = _BODY_FONT
        worksheet.cell(offset, 3, value).font = _BODY_FONT

    note_row = len(rows) + 6
    for line in (
        "How to use this file:",
        "1. Open the 'New Masterlist SKUs' tab.",
        "2. For each row, either paste a SellUp SKU ID into the green",
        "   'Link to SellUp SKU ID' column, or pick a 'Reviewer Decision'.",
        "3. Save, then upload this file back into the app as the SKU Registry.",
        "Suggested SellUp SKU ID shows the tool's best guess — copy it across",
        "if it looks right. Rows already linked are on the Locked Matches tab.",
    ):
        cell = worksheet.cell(note_row, 2, line)
        cell.font = Font(size=9, italic=not line.endswith(":"), bold=line.endswith(":"))
        note_row += 1

    stamp = worksheet.cell(note_row + 1, 2, f"Generated {generated}")
    stamp.font = Font(size=9, italic=True)


def _locked_rows(locked: list[LockedMatch]) -> list[list[object]]:
    return [
        [
            idx,
            m.sellup.sheet,
            m.sellup.sku_id,
            m.sellup.model,
            m.sellup.storage_label,
            m.sellup.connectivity_label,
            m.sellup.colour,
            m.slot,
            m.masterlist_ids,
            m.masterlist_categories,
            m.masterlist_labels,
            m.available_quantities,
            m.target_stock,
            len(m.pos_rows),
            m.origin,
        ]
        for idx, m in enumerate(locked, start=1)
    ]


def _suggestion_cells(item: NewMasterlistSku) -> list[object]:
    """The five suggestion columns for one review row."""
    ranked = list(item.suggestions or [])
    best = ranked[0] if ranked else None

    def label(index: int) -> str:
        if len(ranked) <= index:
            return ""
        alt = ranked[index]
        return f"{alt.sellup.sku_id} · {alt.sellup.display} [{alt.score}]"

    return [
        best.sellup.sku_id if best else "",
        best.sellup.display if best else "",
        best.confidence if best else "no suggestion",
        best.score if best else "",
        ", ".join(best.reasons) if best else "",
        label(1),
        label(2),
    ]


def _new_sku_rows(new_skus: list[NewMasterlistSku]) -> list[list[object]]:
    rows: list[list[object]] = []
    for idx, item in enumerate(new_skus, start=1):
        rows.append(
            [
                idx,
                item.pos.stock_type_id,
                item.pos.category,
                item.pos.brand,
                item.pos.model,
                item.pos.colour,
                item.pos.available_qty,
                item.pos.slot,
                *_suggestion_cells(item),
                item.linked_sku_id,
                item.decision,
                item.notes,
            ]
        )
    return rows


def _match_review_rows(rows) -> list[list[object]]:
    return [
        [
            idx,
            row.sheet,
            row.sku_id,
            row.model,
            row.storage_label,
            row.connectivity_label,
            row.colour,
            "",
            row.current_qty.get(config.SLOT_NEW_NA, ""),
            "",
            "",
            "no POS source recorded",
        ]
        for idx, row in enumerate(rows, start=1)
    ]


def _unsold_rows(pos_rows) -> list[list[object]]:
    return [
        [
            idx,
            r.stock_type_id,
            r.category,
            r.brand,
            r.model,
            r.colour,
            r.available_qty,
        ]
        for idx, r in enumerate(pos_rows, start=1)
    ]


def _link_history_rows(result: PipelineResult) -> list[list[object]]:
    """Every confirmed link, whether or not it produced stock today."""
    synced = {m.sellup.sku_id for m in result.locked}
    names = {m.sellup.sku_id: m.sellup.display for m in result.locked}

    rows: list[list[object]] = []
    for sku_id in sorted(result.all_links):
        rows.append(
            [
                sku_id,
                ", ".join(result.all_links[sku_id]),
                names.get(sku_id, ""),
                "synced" if sku_id in synced else "no POS stock this run",
            ]
        )
    for sku_id in sorted(result.no_pos_source - set(result.all_links)):
        rows.append([sku_id, "", "", "no POS source"])
    return rows


def build_registry_workbook(result: PipelineResult, generated: str) -> bytes:
    """Produce the registry workbook as bytes."""
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    sheets: list[tuple[str, Sequence[str], list[list[object]]]] = [
        (config.SHEET_NEW_SKUS, config.NEW_SKUS_HEADERS, _new_sku_rows(result.new_skus)),
        (config.SHEET_LOCKED, config.LOCKED_HEADERS, _locked_rows(result.locked)),
        (
            config.SHEET_MATCH_REVIEW,
            config.MATCH_REVIEW_HEADERS,
            _match_review_rows(result.match_review),
        ),
        (
            config.SHEET_NOT_SELLING,
            config.UNSOLD_HEADERS,
            _unsold_rows(result.not_selling),
        ),
        (config.SHEET_NOT_YET, config.UNSOLD_HEADERS, _unsold_rows(result.not_yet)),
        (
            config.SHEET_LINK_HISTORY,
            config.LINK_HISTORY_HEADERS,
            _link_history_rows(result),
        ),
    ]

    for name, headers, rows in sheets:
        worksheet = workbook.create_sheet(name)
        _write_header(worksheet, headers)
        written = _write_rows(worksheet, rows, len(headers))

        # Tint the reviewer's input cells so they are obvious in Excel.
        if name == config.SHEET_NEW_SKUS and written:
            for header in config.NEW_SKUS_INPUT_COLUMNS:
                col = headers.index(header) + 1
                for row in range(2, written + 2):
                    worksheet.cell(row, col).fill = _GREEN_FILL

    # Dropdown for the reviewer decision column so the file round-trips cleanly.
    worksheet = workbook[config.SHEET_NEW_SKUS]
    decision_col = config.NEW_SKUS_HEADERS.index("Reviewer Decision") + 1
    letter = get_column_letter(decision_col)
    options = ",".join(d for d in config.DECISION_OPTIONS if d)
    validation = DataValidation(
        type="list", formula1=f'"{options}"', allow_blank=True, showDropDown=False
    )
    worksheet.add_data_validation(validation)
    validation.add(f"{letter}2:{letter}5000")

    _summary_sheet(workbook, result, generated)

    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()
