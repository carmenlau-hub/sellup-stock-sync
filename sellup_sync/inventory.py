"""Reader and writer for the SellUp bulk inventory template.

The writing half is deliberately paranoid. The generated file has to be
uploadable to SellUp untouched, so the writer:

* edits the original uploaded bytes at the XML level rather than re-saving,
* only ever assigns to columns G, I and K on data rows,
* refuses to touch a cell in any other column via :func:`_assert_writable`,
* leaves prices, formulas, styling, merged cells and sheet order alone.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from typing import IO

import openpyxl

from . import config
from .normalize import (
    DeviceSpec,
    clean,
    colour_key,
    maker,
    parse_sellup_specs,
    strip_family_prefix,
)
from .xlsx_editor import cell_ref, compare_archives, write_cells


class InventoryParseError(Exception):
    """Raised when the SellUp template does not have the expected shape."""


_COLUMN_TO_SLOT: dict[int, str] = {v: k for k, v in config.SLOT_TO_COLUMN.items()}

# The price cell that sits immediately left of each quantity cell.
SLOT_TO_PRICE_COLUMN: dict[str, int] = {
    slot: column - 1 for slot, column in config.SLOT_TO_COLUMN.items()
}


def _positive(value: object) -> bool:
    """True when a price or quantity cell carries a number above zero."""
    if value is None or value == "":
        return False
    try:
        return float(str(value).strip()) > 0
    except (TypeError, ValueError):
        return False


@dataclass
class SellUpRow:
    """One listing row from the SellUp bulk inventory template."""

    sheet: str
    excel_row: int
    sku_id: str
    brand: str
    model: str
    specs: str
    colour: str
    spec: DeviceSpec
    current_qty: dict[str, object] = field(default_factory=dict)
    current_price: dict[str, object] = field(default_factory=dict)

    @property
    def listed_conditions(self) -> list[str]:
        """Conditions this listing is actually live in.

        A SellUp row exists in the catalogue whether or not the dealer
        sells it. Setting a price is what makes it a real listing, so a
        priced condition counts as live even at zero stock -- that is the
        Honor X7a case: $118 on the shelf with nothing feeding it.
        """
        live: list[str] = []
        for slot in config.ALL_SLOTS:
            if _positive(self.current_price.get(slot)) or _positive(
                self.current_qty.get(slot)
            ):
                live.append(slot)
        return live

    @property
    def is_listed(self) -> bool:
        return bool(self.listed_conditions)

    @property
    def maker(self) -> str:
        return maker(self.brand)

    @property
    def storage_label(self) -> str:
        return self.spec.storage_label

    @property
    def connectivity_label(self) -> str:
        return self.spec.connectivity_label

    @property
    def display(self) -> str:
        """``'iPhone 17 Pro Max ; 256GB ; Silver'`` style label."""
        parts = [self.model, self.specs, self.colour]
        return " ; ".join(p for p in parts if p)

    def match_key(self) -> tuple:
        return (self.maker, *self.spec.identity(), colour_key(self.colour))

    def loose_key(self) -> tuple:
        return (self.maker, *self.spec.loose_identity(), colour_key(self.colour))

    def family_key(self) -> tuple:
        base = strip_family_prefix(self.spec.base).replace(" ", "")
        return (
            self.maker, base, self.spec.storage_gb, self.spec.network,
            self.spec.case_size_mm, colour_key(self.colour),
        )


@dataclass
class SellUpInventory:
    """Every listing row across the four SellUp worksheets."""

    rows: list[SellUpRow]
    sheet_names: list[str]
    source_bytes: bytes

    def by_sku(self) -> dict[str, SellUpRow]:
        """SKU ID to row. Duplicate SKU IDs keep their first occurrence."""
        out: dict[str, SellUpRow] = {}
        for row in self.rows:
            out.setdefault(row.sku_id, row)
        return out

    def cell_value(self, sheet: str, excel_row: int, column: int) -> object:
        """Current value of a quantity cell, used to skip no-op writes."""
        slot = _COLUMN_TO_SLOT.get(column)
        if slot is None:
            return None
        row = self._row_index().get((sheet, excel_row))
        return row.current_qty.get(slot) if row else None

    def _row_index(self) -> dict[tuple[str, int], SellUpRow]:
        if not hasattr(self, "_index_cache"):
            object.__setattr__(
                self, "_index_cache", {(r.sheet, r.excel_row): r for r in self.rows}
            )
        return self._index_cache  # type: ignore[attr-defined]

    def duplicate_skus(self) -> dict[str, int]:
        counts: dict[str, int] = {}
        for row in self.rows:
            counts[row.sku_id] = counts.get(row.sku_id, 0) + 1
        return {k: v for k, v in counts.items() if v > 1}


def _normalise_header(value: object) -> str:
    """Headers contain embedded newlines; flatten them for comparison."""
    return clean(str(value or "").replace("\n", " ")).upper()


def validate_inventory_workbook(workbook) -> list[str]:
    """Return human-readable problems with the SellUp template layout."""
    problems: list[str] = []

    missing = [s for s in config.SELLUP_SHEETS if s not in workbook.sheetnames]
    if missing:
        problems.append(
            "SellUp inventory file is missing required worksheet(s): "
            + ", ".join(missing)
        )

    for sheet_name in config.SELLUP_SHEETS:
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        for idx, expected in enumerate(config.EXPECTED_SELLUP_HEADERS, start=1):
            actual = _normalise_header(worksheet.cell(config.SELLUP_HEADER_ROW, idx).value)
            if actual != expected.upper():
                problems.append(
                    f"'{sheet_name}' row {config.SELLUP_HEADER_ROW} column {idx}: "
                    f"expected '{expected}', found '{actual or '(blank)'}'. "
                    "The template layout has changed -- writing stock would put "
                    "quantities in the wrong columns."
                )
        if worksheet.max_row < config.SELLUP_FIRST_DATA_ROW:
            problems.append(f"'{sheet_name}' contains no listing rows.")

    return problems


def load_inventory(source: str | IO[bytes] | bytes) -> SellUpInventory:
    """Read the SellUp template, retaining the original bytes for writing."""
    if isinstance(source, bytes):
        raw = source
    elif hasattr(source, "read"):
        if hasattr(source, "seek"):
            source.seek(0)
        raw = source.read()
    else:
        with open(source, "rb") as handle:
            raw = handle.read()

    workbook = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    try:
        problems = validate_inventory_workbook(workbook)
        if problems:
            raise InventoryParseError("\n".join(problems))

        rows: list[SellUpRow] = []
        for sheet_name in config.SELLUP_SHEETS:
            worksheet = workbook[sheet_name]
            for excel_row in range(config.SELLUP_FIRST_DATA_ROW, worksheet.max_row + 1):
                sku_id = clean(worksheet.cell(excel_row, config.COL_SKU_ID).value)
                if not sku_id:
                    continue
                rows.append(
                    SellUpRow(
                        sheet=sheet_name,
                        excel_row=excel_row,
                        sku_id=sku_id,
                        brand=clean(worksheet.cell(excel_row, config.COL_BRAND).value),
                        model=clean(worksheet.cell(excel_row, config.COL_MODEL).value),
                        specs=clean(worksheet.cell(excel_row, config.COL_SPECS).value),
                        colour=clean(worksheet.cell(excel_row, config.COL_COLOR).value),
                        spec=parse_sellup_specs(
                            worksheet.cell(excel_row, config.COL_MODEL).value,
                            worksheet.cell(excel_row, config.COL_SPECS).value,
                        ),
                        current_qty={
                            slot: worksheet.cell(excel_row, col).value
                            for slot, col in config.SLOT_TO_COLUMN.items()
                        },
                        current_price={
                            slot: worksheet.cell(excel_row, col).value
                            for slot, col in SLOT_TO_PRICE_COLUMN.items()
                        },
                    )
                )

        return SellUpInventory(
            rows=rows, sheet_names=list(workbook.sheetnames), source_bytes=raw
        )
    finally:
        workbook.close()


def _assert_writable(column: int) -> None:
    """Guard against any code path attempting to write outside G / I / K."""
    if column not in config.WRITABLE_COLUMNS:
        raise InventoryParseError(
            f"Refusing to write to column {column}: only columns "
            f"{sorted(config.WRITABLE_COLUMNS)} (G, I, K) may be modified."
        )


@dataclass
class QuantityAssignment:
    """A single quantity destined for one cell of the SellUp template."""

    sheet: str
    excel_row: int
    slot: str
    quantity: int
    sku_id: str = ""

    @property
    def column(self) -> int:
        return config.SLOT_TO_COLUMN[self.slot]


@dataclass
class WriteReport:
    """What the writer actually changed."""

    cells_written: int = 0
    cells_unchanged: int = 0
    per_slot: dict[str, int] = field(default_factory=dict)
    per_sheet: dict[str, int] = field(default_factory=dict)


def _as_int(value: object) -> int | None:
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def write_quantities(
    inventory: SellUpInventory,
    assignments: list[QuantityAssignment],
) -> tuple[bytes, WriteReport]:
    """Apply quantities to the original SellUp workbook."""
    report = WriteReport()
    edits: dict[str, dict[str, int]] = {}

    for assignment in assignments:
        column = assignment.column
        _assert_writable(column)

        if assignment.excel_row < config.SELLUP_FIRST_DATA_ROW:
            raise InventoryParseError(
                f"Refusing to write to row {assignment.excel_row}: data starts "
                f"at row {config.SELLUP_FIRST_DATA_ROW}."
            )

        current = inventory.cell_value(assignment.sheet, assignment.excel_row, column)
        if current is not None and _as_int(current) == int(assignment.quantity):
            report.cells_unchanged += 1
            continue

        ref = cell_ref(assignment.excel_row, column)
        edits.setdefault(assignment.sheet, {})[ref] = int(assignment.quantity)

        report.cells_written += 1
        report.per_slot[assignment.slot] = report.per_slot.get(assignment.slot, 0) + 1
        report.per_sheet[assignment.sheet] = report.per_sheet.get(assignment.sheet, 0) + 1

    produced, _ = write_cells(inventory.source_bytes, edits)
    return produced, report


def diff_against_source(original: bytes, produced: bytes) -> list[str]:
    """Verify that only columns G, I and K differ between two workbooks."""
    violations: list[str] = []

    for part in compare_archives(original, produced):
        if not part.startswith("xl/worksheets/"):
            violations.append(
                f"Zip entry '{part}' was modified. Only worksheet parts may change."
            )

    wb_a = openpyxl.load_workbook(io.BytesIO(original))
    wb_b = openpyxl.load_workbook(io.BytesIO(produced))
    try:
        if wb_a.sheetnames != wb_b.sheetnames:
            violations.append(
                f"Worksheet list changed: {wb_a.sheetnames} -> {wb_b.sheetnames}"
            )
            return violations

        for sheet_name in wb_a.sheetnames:
            ws_a, ws_b = wb_a[sheet_name], wb_b[sheet_name]
            if ws_a.max_row != ws_b.max_row or ws_a.max_column != ws_b.max_column:
                violations.append(
                    f"'{sheet_name}' dimensions changed: "
                    f"{ws_a.max_row}x{ws_a.max_column} -> {ws_b.max_row}x{ws_b.max_column}"
                )
                continue

            for row in range(1, ws_a.max_row + 1):
                for col in range(1, ws_a.max_column + 1):
                    va = ws_a.cell(row, col).value
                    vb = ws_b.cell(row, col).value
                    if va == vb:
                        continue
                    if col in config.WRITABLE_COLUMNS and row >= config.SELLUP_FIRST_DATA_ROW:
                        continue
                    violations.append(
                        f"'{sheet_name}' cell "
                        f"{openpyxl.utils.get_column_letter(col)}{row} changed: "
                        f"{va!r} -> {vb!r}"
                    )
                    if len(violations) >= 25:
                        violations.append("... further differences suppressed.")
                        return violations
        return violations
    finally:
        wb_a.close()
        wb_b.close()
