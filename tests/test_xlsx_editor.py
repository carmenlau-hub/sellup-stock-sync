"""Tests for the surgical XLSX editor.

These build a small workbook in memory, edit it, and assert that nothing
outside the target cells moved -- including at the zip-entry level.
"""

from __future__ import annotations

import io
import sys
import zipfile
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from sellup_sync.xlsx_editor import (  # noqa: E402
    XlsxEditError,
    cell_ref,
    compare_archives,
    write_cells,
)


@pytest.fixture()
def workbook_bytes() -> bytes:
    """A two-sheet workbook shaped like the SellUp template."""
    workbook = openpyxl.Workbook()
    first = workbook.active
    first.title = "Smartphones"
    first["A1"] = "DEALER INVENTORY BULK UPDATE"
    first["A3"] = "SKU ID"
    for row in range(4, 12):
        first.cell(row, 1, f"SKU-{row:06d}")
        first.cell(row, 3, "Galaxy S25")
        first.cell(row, 6, 599)   # F price
        first.cell(row, 7, "")    # G qty
        first.cell(row, 9, "")    # I qty
        first.cell(row, 11, "")   # K qty

    second = workbook.create_sheet("Tablets")
    second["A3"] = "SKU ID"
    for row in range(4, 8):
        second.cell(row, 1, f"TAB-{row:06d}")
        second.cell(row, 7, "")

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_cell_ref():
    assert cell_ref(4, 7) == "G4"
    assert cell_ref(4, 9) == "I4"
    assert cell_ref(4, 11) == "K4"
    assert cell_ref(100, 27) == "AA100"


def test_writes_the_requested_values(workbook_bytes):
    produced, report = write_cells(
        workbook_bytes, {"Smartphones": {"G4": 21, "I4": 1, "K4": 3}}
    )
    assert report.cells_written == 3

    sheet = openpyxl.load_workbook(io.BytesIO(produced))["Smartphones"]
    assert sheet["G4"].value == 21
    assert sheet["I4"].value == 1
    assert sheet["K4"].value == 3


def test_values_are_written_as_numbers(workbook_bytes):
    """A Qty written as text would be rejected by the SellUp importer."""
    produced, _ = write_cells(workbook_bytes, {"Smartphones": {"G4": 21}})
    value = openpyxl.load_workbook(io.BytesIO(produced))["Smartphones"]["G4"].value
    assert isinstance(value, int)
    assert not isinstance(value, str)


def test_untouched_cells_keep_their_values(workbook_bytes):
    produced, _ = write_cells(workbook_bytes, {"Smartphones": {"G4": 21}})
    before = openpyxl.load_workbook(io.BytesIO(workbook_bytes))["Smartphones"]
    after = openpyxl.load_workbook(io.BytesIO(produced))["Smartphones"]

    for row in range(1, before.max_row + 1):
        for col in range(1, before.max_column + 1):
            if (row, col) == (4, 7):
                continue
            assert before.cell(row, col).value == after.cell(row, col).value, (
                f"cell {cell_ref(row, col)} changed"
            )


def test_only_edited_sheet_parts_differ(workbook_bytes):
    """sharedStrings, styles and every other part must be byte-identical."""
    produced, _ = write_cells(workbook_bytes, {"Smartphones": {"G4": 21}})
    changed = compare_archives(workbook_bytes, produced)
    assert changed == ["xl/worksheets/sheet1.xml"]


def test_other_sheets_are_byte_identical(workbook_bytes):
    produced, _ = write_cells(workbook_bytes, {"Smartphones": {"G4": 21}})
    with zipfile.ZipFile(io.BytesIO(workbook_bytes)) as a, \
         zipfile.ZipFile(io.BytesIO(produced)) as b:
        assert a.read("xl/worksheets/sheet2.xml") == b.read("xl/worksheets/sheet2.xml")


def test_zip_entry_list_is_preserved(workbook_bytes):
    produced, _ = write_cells(workbook_bytes, {"Smartphones": {"G4": 21}})
    with zipfile.ZipFile(io.BytesIO(workbook_bytes)) as a, \
         zipfile.ZipFile(io.BytesIO(produced)) as b:
        assert [i.filename for i in a.infolist()] == [i.filename for i in b.infolist()]


def test_editing_is_idempotent(workbook_bytes):
    once, _ = write_cells(workbook_bytes, {"Smartphones": {"G4": 21}})
    twice, _ = write_cells(once, {"Smartphones": {"G4": 21}})
    assert once == twice


def test_zero_is_written_not_skipped(workbook_bytes):
    """0 delists a listing, so it must reach the cell as a real zero."""
    produced, _ = write_cells(workbook_bytes, {"Smartphones": {"G4": 0}})
    assert openpyxl.load_workbook(io.BytesIO(produced))["Smartphones"]["G4"].value == 0


def test_unknown_sheet_raises(workbook_bytes):
    with pytest.raises(XlsxEditError, match="not found"):
        write_cells(workbook_bytes, {"Laptops": {"G4": 1}})


def test_missing_cell_raises(workbook_bytes):
    """A reference that does not exist means the row index was miscomputed."""
    with pytest.raises(XlsxEditError, match="do not exist"):
        write_cells(workbook_bytes, {"Smartphones": {"G9999": 1}})


def test_multiple_sheets_edited_together(workbook_bytes):
    produced, report = write_cells(
        workbook_bytes,
        {"Smartphones": {"G4": 5, "G5": 6}, "Tablets": {"G4": 7}},
    )
    assert report.cells_written == 3
    workbook = openpyxl.load_workbook(io.BytesIO(produced))
    assert workbook["Smartphones"]["G4"].value == 5
    assert workbook["Smartphones"]["G5"].value == 6
    assert workbook["Tablets"]["G4"].value == 7
